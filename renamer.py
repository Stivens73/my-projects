import re
from openpyxl import load_workbook
#import xlsxwriter

def is_match(word):

    for m in maps:
        pattern = re.compile("(" + m[0] + ")" )
        #print(m[0])
        result = re.search(pattern, word)
        if result:
            #print(word)
            word = re.sub(m[0], m[1], word)
            print(word)

    return word

def read_dict():

    map_of_cases = []

    for line in open("dict.txt", encoding="utf-8"):

        first = line[0:line.find(';')]
        if first.find('\\.'):

            first = first.replace('.', "\\.", 10)
            #print(first)
        second = line[line.find(';') + 1:len(line) - 1]

        x = [first, second]

        map_of_cases.append(x)

    return map_of_cases

def read_price():
    f = open("output.txt", 'w', encoding="utf-8")
    for line in open("file.txt", encoding="utf-8"):

        buf = is_match(line)
        f.write(buf)



#file = open('file.txt', "rb")
maps = read_dict()


wb2 = load_workbook('price.xlsx')

#workbook = xlsxwriter.Workbook('price.xlsx')



maps.sort(key=lambda t : tuple(t[0].lower()))
f = open("sorted_dict.txt", 'w', encoding="utf-8")
for m in maps:
    f.write(m[0] + ";" + m[1] + "\n")


read_price()
#wb2.save("price.xlsx")

#for m in maps:
#    print(m);